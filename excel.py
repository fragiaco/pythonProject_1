import openpyxl
wb = openpyxl.Workbook()
wb.save('our first notebook.xlsx')
wb = openpyxl.load_workbook('our first notebook.xlsx')
ws = wb.create_sheet('new_Sheet', 0)
ws2 = wb.create_sheet('new_Sheet2', 2)
wb.save('our first notebook.xlsx')
for sheet in wb:
    print(sheet.title)
